import openpyxl



#Считывание данных
def data(file, n, row, column):
    marks = []
    count = 0
    # Расположение файла
    wb = openpyxl.load_workbook(file, data_only=True)
    sheets = wb.sheetnames
    my_sheet_obj = wb[sheets[n-1]]
    #Сбор информации
    my_row = row
    my_column = column
    flag = True
    while flag == True:
        for i in range(my_row, 1000):
            cell_obj = my_sheet_obj.cell(row = i, column = my_column)
            if cell_obj.value is not None:
                count += 1
                my_row += 1
                marks.append(cell_obj.value)
            else:
                flag = False
                break
    print("Оценки: ", marks)
    return marks, count