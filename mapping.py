import openpyxl


def connection(file1, n1, row1, column1, column2):
    # Расположение файла 1
    wb1 = openpyxl.load_workbook(file1, data_only=True)
    sheets1 = wb1.sheetnames
    my_sheet_obj1 = wb1[sheets1[n1 - 1]]
    # Расположение файла 2
    wb2 = openpyxl.load_workbook("проекты итоговая ведомость_20-24 апреля (1).xlsx", data_only=True)
    sheets2 = wb2.sheetnames
    my_sheet_obj2 = []
    for sheet in range(10):
        my_sheet_obj2.append(wb2[sheets2[sheet]])

    # Сбор информации о проектах
    count = 0
    marks = []
    dic = {}
    dic3 = {}
    my_column = column2
    rows = [40, 18, 18, 18, 18, 18, 16, 16, 16, 16]
    for r in rows:
        for i in range(r, 1000):
            cell_obj_id = my_sheet_obj2[count].cell(row=i, column=2)
            cell_obj_name = my_sheet_obj2[count].cell(row=i, column=4)
            cell_obj_hour = my_sheet_obj2[count].cell(row=i, column=6)
            cell_obj_mark = my_sheet_obj2[count].cell(row=i, column=my_column)
            if cell_obj_id.value is not None:
                marks.append(cell_obj_mark.value)
                if type(cell_obj_hour.value) is str:
                    cell_obj_hour.value = cell_obj_hour.value.replace(',', '.')
                    cell_obj_hour.value = float(cell_obj_hour.value)
                if None or cell_obj_hour.value == 0:
                    cell_obj_hour.value = None
                dic[cell_obj_name.value] = marks[0]
                dic3[cell_obj_name.value] = [cell_obj_hour.value, cell_obj_mark.value]
                marks = []
            else:
                break
        count += 1

    # Сбор информации о успеваемости
    my_row = row1
    my_column = column1
    dic2 = {}
    marks = []
    flag = True
    while flag:
        for i in range(my_row, 1000):
            cell_obj_name = my_sheet_obj1.cell(row=i, column=3)
            cell_obj_mark = my_sheet_obj1.cell(row=i, column=my_column)
            if cell_obj_name.value is not None:
                for key in dic.keys():
                    if key == cell_obj_name.value:
                        marks.append(dic[key])
                        if type(cell_obj_mark.value) is str:
                            cell_obj_mark.value = cell_obj_mark.value.replace(',', '.')
                            marks.append(float(cell_obj_mark.value))
                        if type(cell_obj_mark.value) is int:
                            marks.append(cell_obj_mark.value)
                        if type(cell_obj_mark.value) is float:
                            marks.append(cell_obj_mark.value)
                        dic2[cell_obj_name.value] = marks
                        marks = []
            else:
                flag = False

    for key in dic2.keys():
        print(key, '->', dic2[key])

    distribution_density(dic2)
    record(file1, dic2)


# Разделение на оценки за проекты и среднии оценки
def distribution_density(dic):
    uni_mark = []
    proj_mark = []
    for items in dic.items():
        marks = items
        uni_mark.append(marks[1][1])
        proj_mark.append(marks[1][0])

    print("Среднии оценки", uni_mark)
    print("Проектные оценки", proj_mark)


def record(file, dic2):
    wb = openpyxl.load_workbook("Общие оценки.xlsx")
    flag = False
    sheets = wb.sheetnames
    for sheet in sheets:
        if sheet.title() == file:
            flag = True
    if flag == False:
        my_sheet = wb.create_sheet()
        file = file.replace(".xlsx", '')
        file = file.replace("Сводная ведомость_", '')
        my_sheet.title = file
        c1 = my_sheet.cell(row=1, column=1)
        c1.value = "ФИО"
        c2 = my_sheet.cell(row=1, column=2)
        c2.value = "Проектная оценка"
        c3 = my_sheet.cell(row=1, column=3)
        c3.value = "Средняя оценка"
        count = 2
        for key in dic2.keys():
            a = my_sheet.cell(row=count, column=1)
            a.value = key
            b = my_sheet.cell(row=count, column=2)
            b.value = dic2[key][0]
            c = my_sheet.cell(row=count, column=3)
            c.value = dic2[key][1]
            count += 1
    else:
        print("Данные уже внесены в таблицу 'Общие оценки'")
    wb.save("Общие оценки.xlsx")
